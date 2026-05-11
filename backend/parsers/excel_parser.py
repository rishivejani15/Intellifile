import zipfile
from openpyxl import load_workbook

def extract_excel_structure(file_path):
    """
    Extracts sheets, cells, and formulas from a .xlsx file.
    Also detects hidden macros.
    """
    has_macros = False
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            if 'xl/vbaProject.bin' in z.namelist():
                has_macros = True
    except Exception:
        pass

    try:
        # Load twice so we can compare both the formula text and the cached display value.
        wb_formulas = load_workbook(file_path, data_only=False)
        wb_values = load_workbook(file_path, data_only=True)

        sheets_data = {}

        for sheet in wb_formulas.sheetnames:
            ws_formulas = wb_formulas[sheet]
            ws_values = wb_values[sheet] if sheet in wb_values.sheetnames else None
            sheet_content = {}

            for row in ws_formulas.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is not None:
                        cached_value = None
                        if ws_values is not None:
                            cached_value = ws_values[cell.coordinate].value

                        displayed_value = cached_value if cached_value is not None else cell.value
                        sheet_content[cell.coordinate] = {
                            "value": str(cell.value),
                            "formula": cell.value if cell.data_type == "f" else None,
                            "cached_value": None if cached_value is None else str(cached_value),
                            "display_value": str(displayed_value)
                        }

            sheets_data[sheet] = sheet_content

        return {
            "sheets": sheets_data,
            "has_macros": has_macros
        }
    except Exception as e:
        print(f"[ExcelParser] Error: {str(e)}")
        return {"sheets": {}, "has_macros": has_macros}